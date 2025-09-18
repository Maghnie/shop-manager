from datetime import datetime, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Tuple, Any
import json

from django.db import models
from django.db.models import (
    Sum, F, DecimalField, ExpressionWrapper, Value, Count, Avg, Q
)
from django.db.models.functions import TruncDate, TruncHour, TruncWeek, TruncMonth, TruncYear, Coalesce
from django.utils import timezone
from django.core.cache import cache
from django.conf import settings

from .models import AnalyticsCache
from sales.models import Sale, SaleItem
from inventory.models import Product, Inventory


class CacheService:
    """Manage analytics caching"""
    
    @staticmethod
    def generate_cache_key(prefix: str, **kwargs) -> str:
        """Generate consistent cache key"""
        key_parts = [prefix]
        for k, v in sorted(kwargs.items()):
            if isinstance(v, datetime):
                v = v.isoformat()
            key_parts.append(f"{k}:{v}")
        return "_".join(key_parts)
    
    @staticmethod
    def get_cached_data(cache_key: str) -> Optional[Dict]:
        """Get data from cache (Redis or database fallback)"""
        # Try Redis cache first if available
        if hasattr(settings, 'CACHES') and 'redis' in str(settings.CACHES.get('default', {}).get('BACKEND', '')).lower():
            data = cache.get(cache_key)
            if data is not None:
                return data
        
        # Fallback to database cache
        try:
            cache_entry = AnalyticsCache.objects.get(cache_key=cache_key)
            if not cache_entry.is_expired:
                return cache_entry.data
            else:
                cache_entry.delete()
        except AnalyticsCache.DoesNotExist:
            pass
        
        return None
    
    @staticmethod
    def set_cached_data(cache_key: str, data: Dict, hours: int = 24):
        """Set data in cache"""
        # Try Redis cache first
        if hasattr(settings, 'CACHES') and 'redis' in str(settings.CACHES.get('default', {}).get('BACKEND', '')).lower():
            cache.set(cache_key, data, hours * 3600)
        
        # Always store in database as backup
        expires_at = timezone.now() + timedelta(hours=hours)
        AnalyticsCache.objects.update_or_create(
            cache_key=cache_key,
            defaults={
                'data': data,
                'expires_at': expires_at
            }
        )
    
    @staticmethod
    def invalidate_analytics_cache():
        """Invalidate all analytics caches when data changes"""
        # Clear Redis cache patterns
        if hasattr(settings, 'CACHES') and 'redis' in str(settings.CACHES.get('default', {}).get('BACKEND', '')).lower():
            cache.clear()
        
        # Clear database cache
        AnalyticsCache.objects.filter(
            cache_key__startswith='analytics_'
        ).delete()


class TimeSeriesService:
    """Handle time series analytics calculations"""
    
    RESOLUTION_FUNCTIONS = {
        'hourly': TruncHour,
        'daily': TruncDate,
        'weekly': TruncWeek,
        'monthly': TruncMonth,
        'yearly': TruncYear,
    }
    
    @classmethod
    def get_time_series_data(
        cls,
        date_from: datetime,
        date_to: datetime,
        resolution: str = 'daily'
    ) -> Dict:
        """
        Get time series data for sales revenue, profit, and costs
        
        Args:
            date_from: Start date
            date_to: End date  
            resolution: Time resolution (hourly, daily, weekly, monthly, yearly)
        
        Returns:
            Dict with time series data
        """
        # Generate cache key
        cache_key = CacheService.generate_cache_key(
            'analytics_timeseries',
            date_from=date_from,
            date_to=date_to,
            resolution=resolution
        )
        
        # Try to get from cache
        cached_data = CacheService.get_cached_data(cache_key)
        if cached_data:
            return cached_data
        
        # Calculate time series data
        trunc_func = cls.RESOLUTION_FUNCTIONS.get(resolution, TruncDate)
        
        # Base queryset - only completed sales in date range
        sales_qs = Sale.objects.filter(
            status='completed',
            sale_date__gte=date_from,
            sale_date__lte=date_to
        ).select_related()
        
        # Aggregate by time period
        time_series = sales_qs.annotate(
            period=trunc_func('sale_date', tzinfo=timezone.get_current_timezone())
        ).values('period').annotate(
            # Revenue calculations
            revenue=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F('items__quantity') * F('items__unit_price'),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    )
                ) - F('discount_amount') + ExpressionWrapper(
                    (F('items__quantity') * F('items__unit_price') - F('discount_amount')) * F('tax_percentage') / Value(100),
                    output_field=DecimalField(max_digits=18, decimal_places=2)
                ),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            ),
            
            # Cost calculations  
            costs=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F('items__quantity') * F('items__product__cost_price'),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    )
                ),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            ),
            
            # Count of sales
            sales_count=Count('id'),
            
        ).annotate(
            # Profit = Revenue - Costs
            profit=F('revenue') - F('costs')
        ).order_by('period')
        
        # Convert to list and format for frontend
        time_series_list = []
        for item in time_series:
            time_series_list.append({
                'period': item['period'].isoformat(),
                'revenue': float(item['revenue'] or 0),
                'costs': float(item['costs'] or 0), 
                'profit': float(item['profit'] or 0),
                'sales_count': item['sales_count'],
                'profit_margin': round(
                    (float(item['profit'] or 0) / float(item['revenue'] or 1)) * 100, 2
                ) if item['revenue'] and float(item['revenue']) > 0 else 0
            })
        
        # Calculate summary statistics
        total_revenue = sum(item['revenue'] for item in time_series_list)
        total_costs = sum(item['costs'] for item in time_series_list)
        total_profit = total_revenue - total_costs
        total_sales = sum(item['sales_count'] for item in time_series_list)
        
        result = {
            'data': time_series_list,
            'summary': {
                'total_revenue': total_revenue,
                'total_costs': total_costs,
                'total_profit': total_profit,
                'total_sales': total_sales,
                'profit_margin': round((total_profit / total_revenue) * 100, 2) if total_revenue > 0 else 0,
                'average_sale_value': round(total_revenue / total_sales, 2) if total_sales > 0 else 0
            },
            'meta': {
                'date_from': date_from.isoformat(),
                'date_to': date_to.isoformat(),
                'resolution': resolution,
                'generated_at': timezone.now().isoformat()
            }
        }
        
        # Cache the result
        CacheService.set_cached_data(cache_key, result, hours=2)
        
        return result


class BreakevenService:
    """Handle breakeven analysis calculations"""
    
    @classmethod
    def get_breakeven_analysis(
        cls,
        product_id: Optional[int] = None,
        fixed_costs: Decimal = Decimal('0'),
        sort_by: str = 'performance',  # 'performance', 'profit', 'revenue'
        sort_order: str = 'desc'
    ) -> Dict:
        """
        Calculate breakeven analysis for products
        
        Args:
            product_id: Specific product ID (None for all products)
            fixed_costs: Fixed costs to include in breakeven calculation
            sort_by: How to sort results
            sort_order: 'asc' or 'desc'
        
        Returns:
            Dict with breakeven analysis data
        """
        # Generate cache key
        cache_key = CacheService.generate_cache_key(
            'analytics_breakeven',
            product_id=product_id,
            fixed_costs=str(fixed_costs),
            sort_by=sort_by,
            sort_order=sort_order
        )
        
        # Try to get from cache
        cached_data = CacheService.get_cached_data(cache_key)
        if cached_data:
            return cached_data
        
        # Base product queryset
        products_qs = Product.objects.select_related('type', 'brand', 'inventory')
        
        if product_id:
            products_qs = products_qs.filter(id=product_id)
        
        # Calculate performance data for each product
        results = []
        
        for product in products_qs:
            # Get sales data for this product
            sales_data = SaleItem.objects.filter(
                product=product,
                sale__status='completed'
            ).aggregate(
                total_quantity_sold=Coalesce(Sum('quantity'), Value(0)),
                total_revenue=Coalesce(
                    Sum(
                        ExpressionWrapper(
                            F('quantity') * F('unit_price'),
                            output_field=DecimalField(max_digits=18, decimal_places=2)
                        )
                    ),
                    Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
                ),
                total_cost=Coalesce(
                    Sum(
                        ExpressionWrapper(
                            F('quantity') * F('product__cost_price'),
                            output_field=DecimalField(max_digits=18, decimal_places=2)
                        )
                    ),
                    Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
                )
            )
            
            # Calculate breakeven metrics
            unit_price = product.selling_price
            unit_cost = product.cost_price  
            unit_profit = unit_price - unit_cost
            
            # Breakeven calculation: Fixed Costs / Unit Profit
            breakeven_units = 0
            if unit_profit > 0:
                breakeven_units = int(fixed_costs / unit_profit)
            
            # Actual performance
            actual_quantity = sales_data['total_quantity_sold'] or 0
            actual_revenue = float(sales_data['total_revenue'] or 0)
            actual_cost = float(sales_data['total_cost'] or 0)
            actual_profit = actual_revenue - actual_cost
            
            # Performance score (actual sales vs breakeven)
            performance_score = 0
            if breakeven_units > 0:
                performance_score = (actual_quantity / breakeven_units) * 100
            elif actual_profit > 0:
                performance_score = 100  # Product is profitable without needing fixed cost coverage
            
            results.append({
                'product_id': product.id,
                'product_name_ar': str(product),
                'type_name_ar': product.type.name_ar if product.type else '',
                'brand_name_ar': product.brand.name_ar if product.brand else '',
                'unit_price': float(unit_price),
                'unit_cost': float(unit_cost),
                'unit_profit': float(unit_profit),
                'profit_margin': round((float(unit_profit) / float(unit_price)) * 100, 2) if unit_price > 0 else 0,
                'breakeven_units': breakeven_units,
                'breakeven_revenue': float(breakeven_units * unit_price),
                'fixed_costs_allocated': float(fixed_costs),
                'actual_performance': {
                    'quantity_sold': actual_quantity,
                    'revenue': actual_revenue,
                    'cost': actual_cost,
                    'profit': actual_profit
                },
                'performance_score': round(performance_score, 2),
                'status': cls._get_performance_status(performance_score, actual_profit),
                'available_stock': product.inventory.quantity_in_stock if hasattr(product, 'inventory') else 0
            })
        
        # Sort results
        sort_key_map = {
            'performance': lambda x: x['performance_score'],
            'profit': lambda x: x['actual_performance']['profit'],
            'revenue': lambda x: x['actual_performance']['revenue']
        }
        
        sort_key = sort_key_map.get(sort_by, sort_key_map['performance'])
        results.sort(key=sort_key, reverse=(sort_order == 'desc'))
        
        # Calculate summary
        total_fixed_costs = len(results) * float(fixed_costs) if results else 0
        total_revenue = sum(item['actual_performance']['revenue'] for item in results)
        total_profit = sum(item['actual_performance']['profit'] for item in results)
        
        high_performers = len([r for r in results if r['performance_score'] >= 100])
        low_performers = len([r for r in results if r['performance_score'] < 50])
        
        result = {
            'products': results,
            'summary': {
                'total_products': len(results),
                'high_performers': high_performers,
                'low_performers': low_performers,
                'total_fixed_costs': total_fixed_costs,
                'total_revenue': total_revenue,
                'total_profit': total_profit,
                'average_performance_score': round(
                    sum(r['performance_score'] for r in results) / len(results), 2
                ) if results else 0
            },
            'meta': {
                'fixed_costs': float(fixed_costs),
                'sort_by': sort_by,
                'sort_order': sort_order,
                'generated_at': timezone.now().isoformat()
            }
        }
        
        # Cache the result
        CacheService.set_cached_data(cache_key, result, hours=6)
        
        return result
    
    @staticmethod
    def _get_performance_status(performance_score: float, actual_profit: float) -> str:
        """Determine performance status based on score and profit"""
        if performance_score >= 100:
            return 'excellent'  # Above breakeven
        elif performance_score >= 75:
            return 'good'  # Close to breakeven
        elif performance_score >= 50:
            return 'moderate'  # Moderate performance
        elif actual_profit > 0:
            return 'profitable'  # Profitable but low volume
        else:
            return 'poor'  # Below expectations


class ExportService:
    """Handle data export functionality"""
    
    @staticmethod
    def export_to_csv(data: List[Dict], filename: str) -> str:
        """Export data to CSV format"""
        import csv
        import io
        
        output = io.StringIO()
        
        if not data:
            return output.getvalue()
        
        # Get headers from first item
        headers = list(data[0].keys())
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        for row in data:
            # Flatten nested dictionaries
            flat_row = ExportService._flatten_dict(row)
            writer.writerow(flat_row)
        
        return output.getvalue()
    
    @staticmethod
    def export_to_xlsx(data: List[Dict], filename: str) -> bytes:
        """Export data to XLSX format"""
        try:
            import pandas as pd
            from io import BytesIO
            
            if not data:
                return b''
            
            # Flatten the data
            flattened_data = [ExportService._flatten_dict(row) for row in data]
            
            # Create DataFrame
            df = pd.DataFrame(flattened_data)
            
            # Create Excel file in memory
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Analytics Data', index=False)
            
            return output.getvalue()
            
        except ImportError:
            # Fallback to simple Excel creation without pandas
            return ExportService._create_simple_xlsx(data)
    
    @staticmethod
    def _flatten_dict(d: Dict, parent_key: str = '', sep: str = '_') -> Dict:
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(ExportService._flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Convert lists to comma-separated strings
                items.append((new_key, ', '.join(str(x) for x in v)))
            else:
                items.append((new_key, v))
        return dict(items)
    
    @staticmethod
    def _create_simple_xlsx(data: List[Dict]) -> bytes:
        """Create simple XLSX without pandas (fallback)"""
        try:
            from openpyxl import Workbook
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analytics Data"
            
            if not data:
                output = BytesIO()
                wb.save(output)
                return output.getvalue()
            
            # Flatten first row to get headers
            flattened_first = ExportService._flatten_dict(data[0])
            headers = list(flattened_first.keys())
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                flattened_row = ExportService._flatten_dict(row_data)
                for col_idx, header in enumerate(headers, 1):
                    value = flattened_row.get(header, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except ImportError:
            # If openpyxl is also not available, return empty bytes
            return b''